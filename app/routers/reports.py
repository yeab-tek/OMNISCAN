
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_current_user, get_db
from app.models.po_record import PORecord
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/po/export", summary="Export PO records as PDF or Excel")
async def export_po_records(
    format: str = Query("excel", enum=["pdf", "excel"]),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    # Fetch all POs
    result = await db.execute(select(PORecord).order_by(PORecord.created_at.desc()))
    pos = result.scalars().all()

    if format == "excel":
        return await _export_excel(pos)
    else:
        return await _export_pdf(pos)


async def _export_excel(pos):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a5c3a")

    headers = [
        "PO Number", "PO Date", "Buyer Name", "Buyer Country",
        "Quality", "Qty (Bags)", "Bag Weight (kg)", "Price/Unit",
        "Currency", "Incoterms", "Shipment Start", "Shipment End",
        "Origin Port", "Destination Port", "Payment Terms",
        "Payment Status", "EUDR Compliant", "Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, po in enumerate(pos, 2):
        ws.cell(row=row, column=1, value=po.po_number)
        ws.cell(row=row, column=2, value=str(po.po_date) if po.po_date else "")
        ws.cell(row=row, column=3, value=po.buyer_name)
        ws.cell(row=row, column=4, value=po.buyer_country)
        ws.cell(row=row, column=5, value=po.quality_description)
        ws.cell(row=row, column=6, value=po.quantity_bags)
        ws.cell(row=row, column=7, value=float(po.bag_weight_kg) if po.bag_weight_kg else "")
        ws.cell(row=row, column=8, value=float(po.price_per_unit) if po.price_per_unit else "")
        ws.cell(row=row, column=9, value=po.price_currency)
        ws.cell(row=row, column=10, value=po.incoterms)
        ws.cell(row=row, column=11, value=str(po.shipment_start) if po.shipment_start else "")
        ws.cell(row=row, column=12, value=str(po.shipment_end) if po.shipment_end else "")
        ws.cell(row=row, column=13, value=po.origin_port)
        ws.cell(row=row, column=14, value=po.destination_port)
        ws.cell(row=row, column=15, value=po.payment_terms)
        ws.cell(row=row, column=16, value=po.payment_status)
        ws.cell(row=row, column=17, value="Yes" if po.eudr_compliant else "No" if po.eudr_compliant is False else "Pending")
        ws.cell(row=row, column=18, value=po.status)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"omniscan_po_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def _export_pdf(pos):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("OmniScan — Purchase Orders Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    # Table data
    headers = ["PO Number", "Buyer", "Quality", "Bags", "Shipment", "Payment", "Status"]
    data = [headers] + [
        [
            po.po_number or "",
            (po.buyer_name or "")[:20],
            (po.quality_description or "")[:25],
            str(po.quantity_bags or ""),
            str(po.shipment_start or ""),
            po.payment_status or "",
            po.status or "",
        ]
        for po in pos
    ]

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5c3a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e5e5")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    filename = f"omniscan_po_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )